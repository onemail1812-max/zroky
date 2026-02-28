import base64
import io
import logging
from typing import Optional, Dict, Any

try:
    from pypdf import PdfReader
except ImportError:
    PdfReader = None

try:
    import docx
except ImportError:
    docx = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

logger = logging.getLogger(__name__)

class FileExtractorService:
    """Service to extract searchable/analyzable text from various file formats."""

    @staticmethod
    def extract_text(content: bytes, mime_type: str) -> str:
        """Main entry point for text extraction."""
        if not content:
            return ""

        mime_type = mime_type.lower()

        try:
            # 1. Plain Text / CSV / JSON
            if "text/plain" in mime_type or "text/csv" in mime_type or "application/json" in mime_type:
                return content.decode("utf-8", errors="replace")

            # 2. PDF
            if "application/pdf" in mime_type:
                return FileExtractorService._extract_from_pdf(content)

            # 3. Word (Docx)
            if "wordprocessingml" in mime_type:
                return FileExtractorService._extract_from_docx(content)

            # 4. Excel
            if "spreadsheetml" in mime_type:
                return FileExtractorService._extract_from_xlsx(content)

            return f"[Binary file of type {mime_type} - Contents not directly readable as text]"

        except Exception as e:
            logger.error(f"Text extraction failed for {mime_type}: {e}")
            return f"[Error extracting text from {mime_type}]"

    @staticmethod
    def _extract_from_pdf(content: bytes) -> str:
        """Extract text from PDF bytes using pypdf."""
        if not PdfReader:
            logger.warning("pypdf not installed, skipping PDF extraction")
            return "[PDF extraction library unavailable]"

        try:
            reader = PdfReader(io.BytesIO(content))
            text = ""
            # Limit to first 10 pages to avoid prompt blowup in chat
            for i in range(min(len(reader.pages), 10)):
                page_text = reader.pages[i].extract_text() or ""
                if page_text:
                    text += f"--- Page {i+1} ---\n{page_text}\n\n"
            
            if len(reader.pages) > 10:
                text += f"\n... [{len(reader.pages)-10} more pages truncated] ..."
                
            return text.strip()
        except Exception as e:
            logger.error(f"pypdf extraction failed: {e}")
            return "[Failed to parse PDF content]"

    @staticmethod
    def _extract_from_docx(content: bytes) -> str:
        """Extract text from DOCX bytes using python-docx."""
        if not docx:
            logger.warning("python-docx not installed, skipping DOCX extraction")
            return "[DOCX extraction library unavailable]"

        try:
            doc = docx.Document(io.BytesIO(content))
            text = ""
            # Limit paragraphs to avoid prompt blowup (~500 paragraphs max)
            for i, para in enumerate(doc.paragraphs):
                if i >= 500:
                    text += "\n... [Remaining document truncated] ..."
                    break
                if para.text.strip():
                    text += para.text + "\n"
            
            return text.strip() or "[Empty DOCX file]"
        except Exception as e:
            logger.error(f"docx extraction failed: {e}")
            return "[Failed to parse DOCX content]"

    @staticmethod
    def _extract_from_xlsx(content: bytes) -> str:
        """Extract text from XLSX bytes using openpyxl."""
        if not openpyxl:
            logger.warning("openpyxl not installed, skipping XLSX extraction")
            return "[XLSX extraction library unavailable]"

        try:
            # Read-only mode is faster and consumes less memory
            wb = openpyxl.load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
            text = ""
            
            # Limit extraction
            max_sheets = 3
            max_rows = 100
            
            for index, sheet_name in enumerate(wb.sheetnames):
                if index >= max_sheets:
                    text += f"\n... [{len(wb.sheetnames) - max_sheets} more sheets truncated] ..."
                    break
                    
                text += f"--- Sheet: {sheet_name} ---\n"
                sheet = wb[sheet_name]
                
                row_count = 0
                for row in sheet.iter_rows(values_only=True):
                    if row_count >= max_rows:
                        text += f"... [Remaining rows in '{sheet_name}' truncated] ...\n\n"
                        break
                        
                    # Filter out purely None rows
                    row_values = [str(cell) for cell in row if cell is not None]
                    if row_values:
                        text += " | ".join(row_values) + "\n"
                        row_count += 1
                
                text += "\n"
                
            wb.close()
            return text.strip() or "[Empty XLSX file]"
        except Exception as e:
            logger.error(f"xlsx extraction failed: {e}")
            return "[Failed to parse XLSX content]"

    @staticmethod
    def is_image(mime_type: str) -> bool:
        """Helper to identify if mime type is an image."""
        return mime_type.lower().startswith("image/")
